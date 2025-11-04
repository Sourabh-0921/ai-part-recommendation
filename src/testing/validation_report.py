"""
Validation report generation module.

Generates comprehensive Excel reports with validation results.
"""

import pandas as pd
from typing import List, Dict, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging

from .backtesting import BacktestResult
from ..models.calibration import (
    compute_reliability_curve,
    compute_ece,
    compute_brier_score,
    bootstrap_multilabel_confidence_intervals,
)

logger = logging.getLogger(__name__)


class ValidationReportGenerator:
    """
    Generate comprehensive validation report in Excel format.
    
    Report includes:
    - Executive summary
    - Overall performance metrics
    - Part-level analysis
    - Vehicle model analysis
    - Cost impact analysis
    - Case studies
    """
    
    def __init__(self, output_path: str):
        """
        Initialize report generator.
        
        Args:
            output_path: Path to save Excel file
        """
        self.output_path = output_path
        logger.info(f"Validation report generator initialized: {output_path}")
    
    def generate_report(
        self,
        backtest_results: List[BacktestResult],
        metrics: Dict[str, Any],
        baseline_metrics: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Generate Excel validation report.
        
        Args:
            backtest_results: Results from backtesting
            metrics: Calculated metrics
            baseline_metrics: Optional baseline metrics for comparison
            
        Returns:
            Path to generated report file
        """
        try:
            logger.info(f"Generating validation report: {self.output_path}")
            
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Sheet 1: Executive Summary
                self._create_summary_sheet(metrics, baseline_metrics, writer)
                
                # Sheet 2: Overall Performance
                self._create_performance_sheet(metrics, writer)
                
                # Sheet 3: Part Category Analysis
                self._create_part_analysis_sheet(backtest_results, writer)
                
                # Sheet 4: Vehicle Model Analysis
                self._create_model_analysis_sheet(backtest_results, writer)
                
                # Sheet 5: Detailed Results
                self._create_detailed_results_sheet(backtest_results, writer)
                
                # Sheet 6: Cost Impact (if cost data available)
                self._create_cost_analysis_sheet(backtest_results, writer)

                # Sheet 7: Calibration
                self._create_calibration_sheet(backtest_results, metrics, writer)

                # Sheet 8: Confidence by Part
                self._create_confidence_by_part_sheet(backtest_results, writer)

                # Sheet 9: Confidence by Vehicle Model
                self._create_confidence_by_model_sheet(backtest_results, writer)
            
            logger.info(f"Validation report saved: {self.output_path}")
            return self.output_path
            
        except Exception as e:
            logger.error(f"Error generating validation report: {e}")
            raise
    
    def _create_summary_sheet(
        self,
        metrics: Dict[str, Any],
        baseline_metrics: Optional[Dict[str, Any]],
        writer: pd.ExcelWriter
    ):
        """Create executive summary sheet."""
        # Optionally compute CIs for F1 (bootstrap on aggregated predictions)
        try:
            from ..models.validation import AccuracyMetrics
            # Not recomputing predictions here; CI will be presented in Calibration sheet
            f1_ci_text = "See Calibration sheet"
        except Exception:
            f1_ci_text = "N/A"

        summary_data = {
            'Metric': [
                'ML Precision',
                'ML Recall',
                'ML F1 Score',
                'PM Precision',
                'PM Recall',
                'PM F1 Score',
                'Precision Improvement',
                'Recall Improvement',
                'F1 Improvement',
                'Sample Size',
                'Average Confidence',
                'F1 95% CI',
            ],
            'Value': [
                f"{metrics['ml_metrics']['precision']:.2%}",
                f"{metrics['ml_metrics']['recall']:.2%}",
                f"{metrics['ml_metrics']['f1_score']:.2%}",
                f"{metrics['pm_metrics']['precision']:.2%}",
                f"{metrics['pm_metrics']['recall']:.2%}",
                f"{metrics['pm_metrics']['f1_score']:.2%}",
                f"{metrics['ml_advantage']['precision_improvement']:.2%}",
                f"{metrics['ml_advantage']['recall_improvement']:.2%}",
                f"{metrics['ml_advantage']['f1_improvement']:.2%}",
                f"{metrics['sample_size']:,}",
                f"{metrics.get('avg_confidence', 0.0):.2%}",
                f1_ci_text,
            ],
            'Status': [
                self._get_status(metrics['ml_metrics']['precision'], 0.70),
                self._get_status(metrics['ml_metrics']['recall'], 0.70),
                self._get_status(metrics['ml_metrics']['f1_score'], 0.70),
                '',
                '',
                '',
                self._get_status(
                    metrics['ml_advantage']['precision_improvement'], 0.15
                ),
                self._get_status(
                    metrics['ml_advantage']['recall_improvement'], 0.15
                ),
                self._get_status(
                    metrics['ml_advantage']['f1_improvement'], 0.15
                ),
                '',
                '',
                '',
            ]
        }
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets['Executive Summary']
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Status column formatting
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
            status_cell = worksheet.cell(row=row_idx, column=3)
            if '✅' in str(status_cell.value):
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif '⚠️' in str(status_cell.value):
                status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif '❌' in str(status_cell.value):
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_performance_sheet(
        self,
        metrics: Dict[str, Any],
        writer: pd.ExcelWriter
    ):
        """Create overall performance metrics sheet."""
        performance_data = {
            'Metric Type': [
                'Precision', 'Recall', 'F1 Score',
                'True Positives', 'False Positives', 'False Negatives'
            ],
            'ML Model': [
                metrics['ml_metrics']['precision'],
                metrics['ml_metrics']['recall'],
                metrics['ml_metrics']['f1_score'],
                metrics['ml_metrics']['true_positives'],
                metrics['ml_metrics']['false_positives'],
                metrics['ml_metrics']['false_negatives'],
            ],
            'PM Schedule': [
                metrics['pm_metrics']['precision'],
                metrics['pm_metrics']['recall'],
                metrics['pm_metrics']['f1_score'],
                metrics['pm_metrics']['true_positives'],
                metrics['pm_metrics']['false_positives'],
                metrics['pm_metrics']['false_negatives'],
            ],
            'Difference': [
                metrics['ml_metrics']['precision'] - metrics['pm_metrics']['precision'],
                metrics['ml_metrics']['recall'] - metrics['pm_metrics']['recall'],
                metrics['ml_metrics']['f1_score'] - metrics['pm_metrics']['f1_score'],
                metrics['ml_metrics']['true_positives'] - metrics['pm_metrics']['true_positives'],
                metrics['ml_metrics']['false_positives'] - metrics['pm_metrics']['false_positives'],
                metrics['ml_metrics']['false_negatives'] - metrics['pm_metrics']['false_negatives'],
            ],
        }
        
        df = pd.DataFrame(performance_data)
        df.to_excel(writer, sheet_name='Overall Performance', index=False)
    
    def _create_part_analysis_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create part category analysis sheet."""
        # Collect all parts
        all_parts = set()
        for result in backtest_results:
            all_parts.update(result.ml_recommendations)
            all_parts.update(result.pm_recommendations)
            all_parts.update(result.actual_replaced)
        
        part_stats = []
        for part_code in sorted(all_parts):
            ml_recommended = sum(
                1 for r in backtest_results if part_code in r.ml_recommendations
            )
            pm_recommended = sum(
                1 for r in backtest_results if part_code in r.pm_recommendations
            )
            actually_replaced = sum(
                1 for r in backtest_results if part_code in r.actual_replaced
            )
            ml_correct = sum(
                1 for r in backtest_results
                if part_code in r.ml_recommendations and part_code in r.actual_replaced
            )
            pm_correct = sum(
                1 for r in backtest_results
                if part_code in r.pm_recommendations and part_code in r.actual_replaced
            )
            
            ml_precision = ml_correct / ml_recommended if ml_recommended > 0 else 0
            ml_recall = ml_correct / actually_replaced if actually_replaced > 0 else 0
            pm_precision = pm_correct / pm_recommended if pm_recommended > 0 else 0
            pm_recall = pm_correct / actually_replaced if actually_replaced > 0 else 0
            
            part_stats.append({
                'Part Code': part_code,
                'ML Recommended': ml_recommended,
                'ML Correct': ml_correct,
                'ML Precision': ml_precision,
                'ML Recall': ml_recall,
                'PM Recommended': pm_recommended,
                'PM Correct': pm_correct,
                'PM Precision': pm_precision,
                'PM Recall': pm_recall,
                'Actually Replaced': actually_replaced,
            })
        
        df = pd.DataFrame(part_stats)
        df.to_excel(writer, sheet_name='Part Analysis', index=False)
    
    def _create_model_analysis_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create vehicle model analysis sheet."""
        from collections import defaultdict
        
        model_stats = defaultdict(lambda: {
            'ml_pred': [], 'pm_pred': [], 'actual': []
        })
        
        for result in backtest_results:
            if result.vehicle_model:
                model_stats[result.vehicle_model]['ml_pred'].append(result.ml_recommendations)
                model_stats[result.vehicle_model]['pm_pred'].append(result.pm_recommendations)
                model_stats[result.vehicle_model]['actual'].append(result.actual_replaced)
        
        analysis_rows = []
        for model, stats in model_stats.items():
            if not stats['ml_pred']:
                continue
            
            from ..models.validation import AccuracyMetrics
            
            ml_metrics = AccuracyMetrics.calculate_multilabel_metrics(
                stats['ml_pred'], stats['actual']
            )
            pm_metrics = AccuracyMetrics.calculate_multilabel_metrics(
                stats['pm_pred'], stats['actual']
            )
            
            analysis_rows.append({
                'Vehicle Model': model,
                'Sample Size': len(stats['ml_pred']),
                'ML Precision': ml_metrics['precision'],
                'ML Recall': ml_metrics['recall'],
                'ML F1': ml_metrics['f1_score'],
                'PM Precision': pm_metrics['precision'],
                'PM Recall': pm_metrics['recall'],
                'PM F1': pm_metrics['f1_score'],
                'F1 Improvement': ml_metrics['f1_score'] - pm_metrics['f1_score'],
            })
        
        df = pd.DataFrame(analysis_rows)
        df.to_excel(writer, sheet_name='Vehicle Model Analysis', index=False)
    
    def _create_detailed_results_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create detailed results sheet."""
        detailed_data = []
        for result in backtest_results:
            detailed_data.append({
                'Vehicle ID': result.vehicle_id,
                'Vehicle Model': result.vehicle_model or '',
                'Service Date': result.service_date,
                'Job Card': result.job_card_number or '',
                'ML Recommendations': ', '.join(result.ml_recommendations),
                'PM Recommendations': ', '.join(result.pm_recommendations),
                'Actual Replaced': ', '.join(result.actual_replaced),
                'ML Confidence': ', '.join([
                    f"{k}={v:.2%}" for k, v in result.ml_confidences.items()
                ]),
            })
        
        df = pd.DataFrame(detailed_data)
        df.to_excel(writer, sheet_name='Detailed Results', index=False)
    
    def _create_cost_analysis_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create cost impact analysis sheet (placeholder)."""
        cost_data = {
            'Analysis': [
                'Total ML Recommendations',
                'Total PM Recommendations',
                'Cost Impact Analysis',
                'Note: Cost analysis requires part cost data'
            ],
            'Value': [
                sum(len(r.ml_recommendations) for r in backtest_results),
                sum(len(r.pm_recommendations) for r in backtest_results),
                'N/A',
                'To be implemented with part cost master data'
            ]
        }
        
        df = pd.DataFrame(cost_data)
        df.to_excel(writer, sheet_name='Cost Impact', index=False)

    def _create_calibration_sheet(
        self,
        backtest_results: List[BacktestResult],
        metrics: Dict[str, Any],
        writer: pd.ExcelWriter
    ):
        """Create calibration sheet with reliability curve, ECE, and Brier score.
        Uses recommended parts confidences as probabilistic predictions and
        whether they were actually replaced as labels.
        """
        # Collect confidences and labels from backtest results
        confidences: List[float] = []
        labels: List[int] = []
        preds: List[List[str]] = []
        actuals: List[List[str]] = []

        for r in backtest_results:
            actual_set = set(r.actual_replaced)
            actuals.append(r.actual_replaced)
            preds.append(r.ml_recommendations)
            for part, conf in (r.ml_confidences or {}).items():
                confidences.append(float(conf))
                labels.append(1 if part in actual_set else 0)

        # Reliability stats
        rel = compute_reliability_curve(confidences, labels, n_bins=10)
        ece = compute_ece(rel['mean_confidence'], rel['accuracy'], rel['count'])
        brier = compute_brier_score(confidences, labels)

        # Bootstrap CIs for multilabel metrics based on full predictions/actuals
        ci = bootstrap_multilabel_confidence_intervals(preds, actuals, n_bootstrap=300, alpha=0.05)

        # Write reliability table
        rel_df = pd.DataFrame({
            'Bin Center': rel['bin_centers'],
            'Mean Confidence': rel['mean_confidence'],
            'Accuracy': rel['accuracy'],
            'Count': rel['count'],
        })
        rel_df.to_excel(writer, sheet_name='Calibration', index=False)

        # Append metrics below the table
        ws = writer.sheets['Calibration']
        start_row = len(rel_df) + 3
        ws.cell(row=start_row, column=1, value='ECE')
        ws.cell(row=start_row, column=2, value=ece)
        ws.cell(row=start_row + 1, column=1, value='Brier Score')
        ws.cell(row=start_row + 1, column=2, value=brier)
        ws.cell(row=start_row + 3, column=1, value='Bootstrap 95% CI (Precision)')
        ws.cell(row=start_row + 3, column=2, value=f"{ci['precision']['low']:.3f} - {ci['precision']['high']:.3f}")
        ws.cell(row=start_row + 4, column=1, value='Bootstrap 95% CI (Recall)')
        ws.cell(row=start_row + 4, column=2, value=f"{ci['recall']['low']:.3f} - {ci['recall']['high']:.3f}")
        ws.cell(row=start_row + 5, column=1, value='Bootstrap 95% CI (F1)')
        ws.cell(row=start_row + 5, column=2, value=f"{ci['f1_score']['low']:.3f} - {ci['f1_score']['high']:.3f}")

    def _create_confidence_by_part_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create part-level confidence distribution sheet."""
        from collections import defaultdict
        import numpy as np
        part_conf: Dict[str, list] = defaultdict(list)
        for r in backtest_results:
            for part, conf in (r.ml_confidences or {}).items():
                part_conf[part].append(float(conf))
        rows = []
        for part, vals in part_conf.items():
            rows.append({
                'Part Code': part,
                'Count': len(vals),
                'Mean Confidence': float(np.mean(vals)) if vals else 0.0,
                'Median Confidence': float(np.median(vals)) if vals else 0.0,
                'Min Confidence': float(np.min(vals)) if vals else 0.0,
                'Max Confidence': float(np.max(vals)) if vals else 0.0,
            })
        df = pd.DataFrame(rows).sort_values(by='Mean Confidence', ascending=False)
        df.to_excel(writer, sheet_name='Confidence by Part', index=False)

    def _create_confidence_by_model_sheet(
        self,
        backtest_results: List[BacktestResult],
        writer: pd.ExcelWriter
    ):
        """Create vehicle model-level confidence distribution sheet."""
        from collections import defaultdict
        import numpy as np
        model_conf: Dict[str, list] = defaultdict(list)
        for r in backtest_results:
            if r.vehicle_model:
                for _, conf in (r.ml_confidences or {}).items():
                    model_conf[r.vehicle_model].append(float(conf))
        rows = []
        for model, vals in model_conf.items():
            rows.append({
                'Vehicle Model': model,
                'Count': len(vals),
                'Mean Confidence': float(np.mean(vals)) if vals else 0.0,
                'Median Confidence': float(np.median(vals)) if vals else 0.0,
                'Min Confidence': float(np.min(vals)) if vals else 0.0,
                'Max Confidence': float(np.max(vals)) if vals else 0.0,
            })
        df = pd.DataFrame(rows).sort_values(by='Mean Confidence', ascending=False)
        df.to_excel(writer, sheet_name='Confidence by Model', index=False)
    
    @staticmethod
    def _get_status(value: float, threshold: float) -> str:
        """Get status indicator based on threshold."""
        if value >= threshold:
            return '✅ PASS'
        elif value >= threshold * 0.9:
            return '⚠️ WARNING'
        else:
            return '❌ FAIL'

